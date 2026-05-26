#!/usr/bin/env python3
"""Build browser-ready policy impact summaries from Beijing air quality data."""

from __future__ import annotations

import csv
import io
import json
import re
import statistics
import zipfile
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
AIR_DIR = ROOT / "data" / "北京空气质量"
POLICY_XLSX = ROOT / "data" / "policies" / "空气污染治理政策汇总.xlsx"
OUTPUT = ROOT / "src" / "policy-impact-summary.json"

FULL_YEAR_RE = re.compile(r"beijing_(20\d{2})0101-(20\d{2})1231\.zip$")
CSV_DATE_RE = re.compile(r"beijing_all_(\d{8})\.csv$")
METRICS = ("PM2.5", "PM10", "AQI")


def parse_number(value: object) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        number = float(text)
    except ValueError:
        return None
    if number < 0:
        return None
    return number


def mean(values: list[float]) -> float | None:
    return statistics.fmean(values) if values else None


def classify_aqi(aqi: float | None) -> str | None:
    if aqi is None:
        return None
    if aqi <= 50:
        return "excellent"
    if aqi <= 100:
        return "good"
    if aqi <= 150:
        return "light"
    if aqi <= 200:
        return "moderate"
    if aqi <= 300:
        return "heavy"
    return "severe"


def parse_daily_csv(text: str) -> dict[str, float | None]:
    rows = csv.DictReader(io.StringIO(text))
    hourly_metric_values: dict[str, list[float]] = {metric: [] for metric in METRICS}

    for row in rows:
        metric = (row.get("type") or "").strip()
        if metric not in hourly_metric_values:
            continue
        station_values = [
            parsed
            for key, value in row.items()
            if key not in {"date", "hour", "type"} and (parsed := parse_number(value)) is not None
        ]
        row_mean = mean(station_values)
        if row_mean is not None:
            hourly_metric_values[metric].append(row_mean)

    return {metric: mean(values) for metric, values in hourly_metric_values.items()}


def iter_daily_csvs_from_zip(zip_path: Path):
    with zipfile.ZipFile(zip_path) as archive:
        names = [
            name
            for name in archive.namelist()
            if CSV_DATE_RE.search(Path(name).name)
        ]
        for name in sorted(names):
            match = CSV_DATE_RE.search(Path(name).name)
            if not match:
                continue
            raw = archive.read(name)
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("gb18030")
            yield match.group(1), text


def build_air_quality_summary():
    daily = []
    yearly_stats: dict[int, dict[str, object]] = {}
    monthly_accumulator: dict[tuple[int, int], dict[str, list[float]]] = defaultdict(lambda: defaultdict(list))

    zip_paths = sorted(
        path
        for path in AIR_DIR.glob("beijing_*.zip")
        if FULL_YEAR_RE.match(path.name) or path.name == "beijing_20260101-20260425.zip"
    )

    for zip_path in zip_paths:
        for yyyymmdd, text in iter_daily_csvs_from_zip(zip_path):
            current_date = datetime.strptime(yyyymmdd, "%Y%m%d").date()
            values = parse_daily_csv(text)
            if not any(values.values()):
                continue

            record = {
                "date": current_date.isoformat(),
                "year": current_date.year,
                "month": current_date.month,
                "pm25": round(values["PM2.5"], 2) if values["PM2.5"] is not None else None,
                "pm10": round(values["PM10"], 2) if values["PM10"] is not None else None,
                "aqi": round(values["AQI"], 2) if values["AQI"] is not None else None,
                "level": classify_aqi(values["AQI"]),
            }
            daily.append(record)

            for metric, source in (("pm25", values["PM2.5"]), ("pm10", values["PM10"]), ("aqi", values["AQI"])):
                if source is not None:
                    monthly_accumulator[(current_date.year, current_date.month)][metric].append(source)

    by_year: dict[int, list[dict[str, object]]] = defaultdict(list)
    for record in daily:
        by_year[record["year"]].append(record)

    for year, records in sorted(by_year.items()):
        metrics = {}
        for metric in ("pm25", "pm10", "aqi"):
            values = [r[metric] for r in records if r[metric] is not None]
            metrics[metric] = round(mean(values), 2) if values else None

        level_counts = {key: 0 for key in ("excellent", "good", "light", "moderate", "heavy", "severe")}
        for record in records:
            if record["level"]:
                level_counts[record["level"]] += 1

        yearly_stats[year] = {
            "year": year,
            "days": len(records),
            **metrics,
            "excellentGoodDays": level_counts["excellent"] + level_counts["good"],
            "pollutedDays": level_counts["light"] + level_counts["moderate"] + level_counts["heavy"] + level_counts["severe"],
            "heavyPollutionDays": level_counts["heavy"] + level_counts["severe"],
            "levelCounts": level_counts,
            "isFullYear": len(records) >= 360,
        }

    monthly = []
    for (year, month), metric_values in sorted(monthly_accumulator.items()):
        monthly.append({
            "year": year,
            "month": month,
            "pm25": round(mean(metric_values["pm25"]), 2) if metric_values["pm25"] else None,
            "pm10": round(mean(metric_values["pm10"]), 2) if metric_values["pm10"] else None,
            "aqi": round(mean(metric_values["aqi"]), 2) if metric_values["aqi"] else None,
        })

    return daily, list(yearly_stats.values()), monthly


def excel_date_to_iso(value: object, fallback_year: int | None) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        return openpyxl.utils.datetime.from_excel(value).date().isoformat()

    text = str(value or "").strip()
    match = re.search(r"(20\d{2})\s*年\s*(\d{1,2})?\s*月?", text)
    if match:
        year = int(match.group(1))
        month = int(match.group(2) or 1)
        return date(year, month, 1).isoformat()
    if fallback_year:
        return date(fallback_year, 1, 1).isoformat()
    return ""


def build_policy_summary():
    workbook = openpyxl.load_workbook(POLICY_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    policies = []

    for row in rows[1:]:
        values = dict(zip(headers, row))
        if not values.get("政策名称"):
            continue
        year = parse_number(values.get("发布年份"))
        year_int = int(year) if year is not None else None
        policies.append({
            "id": int(parse_number(values.get("序号")) or len(policies) + 1),
            "name": str(values.get("政策名称")).strip("《》"),
            "date": excel_date_to_iso(values.get("发布时间"), year_int),
            "year": year_int,
            "agency": values.get("发布机构") or "",
            "summary": values.get("主要内容") or "",
            "measures": values.get("主要措施") or "",
            "pollutants": values.get("涉及污染物") or "",
            "effect": values.get("实施效果或后续评估信息") or "",
            "url": values.get("政策链接") or "",
        })
    return policies


def get_year(yearly: list[dict[str, object]], year: int):
    return next((item for item in yearly if item["year"] == year), None)


def build_policy_windows(yearly: list[dict[str, object]]):
    windows = [
        ("2013-2017 清洁空气行动计划", 2014, 2017, "压煤、控车、治污、降尘成为主线"),
        ("2018-2020 蓝天保卫战", 2018, 2020, "柴油货车、扬尘、VOCs 被集中治理"),
        ("2021-2025 十四五协同治理", 2021, 2025, "PM2.5 与臭氧协同，治理进入精细化阶段"),
        ("2025 0.1 微克攻坚", 2024, 2025, "从显著压降转向小幅精细改善"),
    ]

    result = []
    for name, start_year, end_year, note in windows:
        start = get_year(yearly, start_year)
        end = get_year(yearly, end_year)
        if not start or not end:
            continue
        start_value = start["pm25"]
        end_value = end["pm25"]
        if start_value is None or end_value is None:
            continue
        result.append({
            "name": name,
            "startYear": start_year,
            "endYear": end_year,
            "startPm25": start_value,
            "endPm25": end_value,
            "deltaPm25": round(end_value - start_value, 2),
            "changePct": round((end_value - start_value) / start_value * 100, 1),
            "startGoodDays": start["excellentGoodDays"],
            "endGoodDays": end["excellentGoodDays"],
            "deltaGoodDays": end["excellentGoodDays"] - start["excellentGoodDays"],
            "note": note,
        })
    return result


def build_findings(yearly: list[dict[str, object]], policy_windows: list[dict[str, object]]):
    base = get_year(yearly, 2014)
    latest_full = max((item for item in yearly if item["isFullYear"]), key=lambda item: item["year"])
    findings = []

    if base and latest_full:
        pm25_change = (latest_full["pm25"] - base["pm25"]) / base["pm25"] * 100
        aqi_change = (latest_full["aqi"] - base["aqi"]) / base["aqi"] * 100
        findings.append({
            "label": "长期改善",
            "value": f"{abs(pm25_change):.1f}%",
            "text": f"2014 到 {latest_full['year']} 年，全市站点均值 PM2.5 从 {base['pm25']} 降至 {latest_full['pm25']} ug/m3，降幅 {abs(pm25_change):.1f}%；AQI 均值同步下降 {abs(aqi_change):.1f}%。",
        })
        findings.append({
            "label": "好天变多",
            "value": f"+{latest_full['excellentGoodDays'] - base['excellentGoodDays']} 天",
            "text": f"AQI 均值 <=100 的天数从 {base['excellentGoodDays']} 天增至 {latest_full['excellentGoodDays']} 天，重污染天数从 {base['heavyPollutionDays']} 天降至 {latest_full['heavyPollutionDays']} 天。",
        })

    steepest = min(policy_windows, key=lambda item: item["changePct"])
    findings.append({
        "label": "政策张力",
        "value": f"{steepest['changePct']}%",
        "text": f"{steepest['name']} 对应的 PM2.5 下降最陡：{steepest['startYear']} 到 {steepest['endYear']} 年下降 {abs(steepest['deltaPm25'])} ug/m3。后期政策仍有效，但边际改善变小。",
    })
    return findings


def main():
    daily, yearly, monthly = build_air_quality_summary()
    full_yearly = [item for item in yearly if item["isFullYear"]]
    policies = build_policy_summary()
    policy_windows = build_policy_windows(yearly)

    payload = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "source": "data/北京空气质量/*.zip + data/policies/空气污染治理政策汇总.xlsx",
            "method": "每日 CSV 先对所有可用监测站求小时均值，再对小时均值求日均值；年度和月度由日均值聚合。",
            "note": "2026 数据截至 2026-04-25，年度趋势图中单独标注为未满年。",
        },
        "yearly": yearly,
        "fullYearly": full_yearly,
        "monthly": monthly,
        "policies": policies,
        "policyWindows": policy_windows,
        "findings": build_findings(yearly, policy_windows),
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
